import os
import re
import json
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests

app = Flask(__name__)
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Control_de_Gastos.xlsx")
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")

def get_google_webhook_url():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("google_webhook_url", "")
        except Exception:
            return ""
    return ""

def set_google_webhook_url(url):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump({"google_webhook_url": url}, f)

# -------------------------------------------------------------
# GENERADOR AUTOMÁTICO DE EXCEL EN LA NUBE SI NO EXISTE
# -------------------------------------------------------------
def asegurar_excel_existente():
    if os.path.exists(EXCEL_PATH):
        return
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Gastos"
    ws.views.sheetView[0].showGridLines = True

    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    CARD_LABEL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    INPUT_CELL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    SUMMARY_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    SUMMARY_RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    TABLE_HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    SUBTITLE_FONT = Font(name="Segoe UI", size=10, italic=True, color="D9E1F2")
    CARD_TITLE_FONT = Font(name="Segoe UI", size=9, bold=True, color="1F4E78")
    CARD_VALUE_FONT = Font(name="Segoe UI", size=14, bold=True, color="000000")
    TABLE_HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
    REGULAR_FONT = Font(name="Segoe UI", size=10)
    
    THIN_SIDE = Side(border_style="thin", color="D9D9D9")
    DOUBLE_BOTTOM = Side(border_style="double", color="1F4E78")
    CARD_BORDER = Side(border_style="thin", color="B4C6E7")
    
    DEFAULT_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    TOTAL_BORDER = Border(top=THIN_SIDE, bottom=DOUBLE_BOTTOM, left=THIN_SIDE, right=THIN_SIDE)

    ws.merge_cells("B2:H2")
    ws["B2"] = "CONTROL DE PRESUPUESTO Y GASTOS PERSONAL"
    ws["B2"].font = TITLE_FONT
    ws["B2"].fill = HEADER_FILL
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws.merge_cells("B3:H3")
    ws["B3"] = "Establece tu monto inicial y registra tus compras para mantener el saldo actualizado automáticamente"
    ws["B3"].font = SUBTITLE_FONT
    ws["B3"].fill = HEADER_FILL
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 18

    ws["B5"] = "MONTO INICIAL"
    ws["B5"].font = CARD_TITLE_FONT
    ws["B5"].fill = CARD_LABEL_FILL
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["B6"] = 5000.00
    ws["B6"].font = CARD_VALUE_FONT
    ws["B6"].fill = INPUT_CELL_FILL
    ws["B6"].number_format = "$#,##0.00"
    ws["B6"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["C5"] = "TOTAL GASTADO"
    ws["C5"].font = CARD_TITLE_FONT
    ws["C5"].fill = CARD_LABEL_FILL
    ws["C5"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["C6"] = "=SUM(F11:F100)"
    ws["C6"].font = CARD_VALUE_FONT
    ws["C6"].fill = SUMMARY_RED_FILL
    ws["C6"].number_format = "$#,##0.00"
    ws["C6"].alignment = Alignment(horizontal="center", vertical="center")

    ws["D5"] = "SALDO DISPONIBLE"
    ws["D5"].font = CARD_TITLE_FONT
    ws["D5"].fill = CARD_LABEL_FILL
    ws["D5"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["D6"] = "=B6-C6"
    ws["D6"].font = CARD_VALUE_FONT
    ws["D6"].fill = SUMMARY_GREEN_FILL
    ws["D6"].number_format = "$#,##0.00"
    ws["D6"].alignment = Alignment(horizontal="center", vertical="center")

    ws["E5"] = "% PRESUPUESTO USADO"
    ws["E5"].font = CARD_TITLE_FONT
    ws["E5"].fill = CARD_LABEL_FILL
    ws["E5"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["E6"] = "=IF(B6>0, C6/B6, 0)"
    ws["E6"].font = CARD_VALUE_FONT
    ws["E6"].fill = CARD_LABEL_FILL
    ws["E6"].number_format = "0.0%"
    ws["E6"].alignment = Alignment(horizontal="center", vertical="center")

    for col in ["B", "C", "D", "E"]:
        ws[f"{col}5"].border = Border(left=CARD_BORDER, right=CARD_BORDER, top=CARD_BORDER, bottom=THIN_SIDE)
        ws[f"{col}6"].border = Border(left=CARD_BORDER, right=CARD_BORDER, top=THIN_SIDE, bottom=CARD_BORDER)

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 28

    headers = [
        ("B10", "Nº"), ("C10", "Fecha"), ("D10", "Categoría"),
        ("E10", "Descripción / Concepto"), ("F10", "Monto ($)"),
        ("G10", "Método de Pago"), ("H10", "Saldo Restante ($)")
    ]
    
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = DEFAULT_BORDER

    ws.row_dimensions[10].height = 24
    ws.column_dimensions["E"].width = 50

    start_row = 11
    total_rows = 50
    for i in range(total_rows):
        current_row = start_row + i
        ws.row_dimensions[current_row].height = 22
        ws[f"B{current_row}"] = i + 1

        if i == 0:
            ws[f"H{current_row}"] = f"=IF(ISBLANK(F{current_row}), B6, B6-F{current_row})"
        else:
            prev_row = current_row - 1
            ws[f"H{current_row}"] = f"=IF(ISBLANK(F{current_row}), \"\", H{prev_row}-F{current_row})"

        ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"E{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"H{current_row}"].alignment = Alignment(horizontal="right", vertical="center")

        ws[f"F{current_row}"].number_format = "$#,##0.00"
        ws[f"H{current_row}"].number_format = "$#,##0.00"

        row_fill = ALT_ROW_FILL if i % 2 == 1 else PatternFill(fill_type=None)
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            cell = ws[f"{col}{current_row}"]
            cell.font = REGULAR_FONT
            cell.border = DEFAULT_BORDER
            if i % 2 == 1:
                cell.fill = row_fill

    summary_row = start_row + total_rows
    ws[f"E{summary_row}"] = "TOTAL GASTADO:"
    ws[f"E{summary_row}"].font = BOLD_FONT
    ws[f"E{summary_row}"].alignment = Alignment(horizontal="right", vertical="center")
    
    ws[f"F{summary_row}"] = f"=SUM(F{start_row}:F{summary_row-1})"
    ws[f"F{summary_row}"].font = BOLD_FONT
    ws[f"F{summary_row}"].number_format = "$#,##0.00"
    ws[f"F{summary_row}"].border = TOTAL_BORDER
    ws[f"F{summary_row}"].fill = CARD_LABEL_FILL

    wb.save(EXCEL_PATH)
    wb.close()

asegurar_excel_existente()

# -------------------------------------------------------------
# LÓGICA DE PROCESAMIENTO DE VOZ EN ESPAÑOL
# -------------------------------------------------------------
CATEGORIAS_KEYWORDS = {
    "Alimentación": ["supermercado", "super", "comida", "cena", "almuerzo", "desayuno", "restaurante", "despensa", "mercado", "café", "pizza", "hamburguesa"],
    "Transporte": ["gasolina", "gas", "uber", "didi", "taxi", "pasaje", "bus", "combustible", "peaje", "estacionamiento", "parqueo"],
    "Vivienda / Servicios": ["luz", "agua", "internet", "wifi", "alquiler", "renta", "teléfono", "celular", "gas domiliciario"],
    "Entretenimiento": ["cine", "película", "juego", "salida", "bar", "cerveza", "fiesta", "concierto", "netflix", "spotify"],
    "Salud": ["farmacia", "doctor", "médico", "medicina", "medicamentos", "consulta", "dentista", "clínica"],
    "Educación": ["curso", "libro", "colegio", "escuela", "universidad"],
    "Ropa y Calzado": ["ropa", "zapatos", "camisa", "pantalón", "tienda"]
}

NUMEROS_TEXTO = {
    "uno": 1, "un": 1, "dos": 2, "tres": 3, "cuatro": 4, "cinco": 5, "seis": 6, "siete": 7, "ocho": 8, "nueve": 9, "diez": 10,
    "quince": 15, "veinte": 20, "treinta": 30, "cuarenta": 40, "cincuenta": 50, "sesenta": 60, "setenta": 70, "ochenta": 80, "noventa": 90,
    "cien": 100, "ciento": 100, "doscientos": 200, "trescientos": 300, "cuatrocientos": 400, "quinientos": 500, "mil": 1000
}

def parse_voice_text(text):
    text_lower = text.lower()
    monto = 0.0
    num_match = re.search(r'\b\d+(?:[\.,]\d{1,2})?\b', text_lower)
    if num_match:
        monto = float(num_match.group(0).replace(',', '.'))
    else:
        words = text_lower.split()
        temp_val = 0
        for w in words:
            if w in NUMEROS_TEXTO:
                temp_val += NUMEROS_TEXTO[w]
        if temp_val > 0:
            monto = float(temp_val)

    categoria_detectada = "Otros"
    for cat, keywords in CATEGORIAS_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            categoria_detectada = cat
            break

    metodo_pago = "Efectivo"
    if any(k in text_lower for k in ["débito", "debito"]):
        metodo_pago = "Tarjeta de Débito"
    elif any(k in text_lower for k in ["crédito", "credito"]):
        metodo_pago = "Tarjeta de Crédito"
    elif any(k in text_lower for k in ["transferencia", "banco", "sinpe", "transferir"]):
        metodo_pago = "Transferencia"

    clean_desc = text
    for remove_word in ["gasté", "gaste", "pagué", "pague", "compré", "compre", "pesos", "quetzales", "dólares", "dolares", "soles", "en", "con", "por"]:
        clean_desc = re.sub(rf'\b{remove_word}\b', '', clean_desc, flags=re.IGNORECASE)
    clean_desc = re.sub(r'\s+', ' ', clean_desc).strip()
    if not clean_desc:
        clean_desc = f"Gasto en {categoria_detectada}"

    return {
        "monto": monto,
        "categoria": categoria_detectada,
        "descripcion": clean_desc.capitalize(),
        "metodo_pago": metodo_pago,
        "raw_text": text
    }

def sync_from_google_sheets():
    google_url = get_google_webhook_url()
    if not google_url:
        return False, "No hay Google Sheets vinculado.", None, None, None, []

    if not google_url.startswith("https://script.google.com"):
        return False, "La URL vinculada debe empezar por 'https://script.google.com/macros/s/.../exec'", None, None, None, []

    try:
        res = requests.get(google_url, allow_redirects=True, timeout=6)
        if res.status_code == 200:
            try:
                data = res.json()
            except Exception:
                return False, "Google devolvió una página de acceso o error HTML. Asegúrate de publicar el Apps Script con Acceso: 'Cualquier persona' (Anyone) y desplegar una NUEVA VERSIÓN.", None, None, None, []

            monto_inicial = float(data.get("monto_inicial", 5000))
            total_gastos = float(data.get("total_gastos", 0))
            saldo_disponible = float(data.get("saldo_disponible", monto_inicial - total_gastos))
            expenses = data.get("expenses", [])

            asegurar_excel_existente()
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Control de Gastos"]
            ws["B6"] = monto_inicial

            for r in range(11, 60):
                ws[f"C{r}"] = None
                ws[f"D{r}"] = None
                ws[f"E{r}"] = None
                ws[f"F{r}"] = None
                ws[f"G{r}"] = None

            for idx, item in enumerate(expenses):
                r = 11 + idx
                if r < 60:
                    ws[f"B{r}"] = idx + 1
                    ws[f"C{r}"] = item.get("fecha", "")
                    ws[f"D{r}"] = item.get("categoria", "Otros")
                    ws[f"E{r}"] = item.get("descripcion", "Gasto")
                    ws[f"F{r}"] = float(item.get("monto", 0))
                    ws[f"G{r}"] = item.get("metodo_pago", "Efectivo")

            wb.save(EXCEL_PATH)
            wb.close()

            return True, f"Sincronizados {len(expenses)} registros desde Google Sheets", monto_inicial, total_gastos, saldo_disponible, expenses
    except Exception as e:
        return False, f"Error de conexión con Google Sheets: {str(e)}", None, None, None, []

    return False, "No se pudo sincronizar", None, None, None, []

def get_excel_summary():
    success, msg, mi, tg, sd, _ = sync_from_google_sheets()
    if success and mi is not None:
        return {"monto_inicial": mi, "total_gastos": tg, "saldo_disponible": sd}

    asegurar_excel_existente()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Control de Gastos"]
    monto_inicial = ws["B6"].value or 0
    total_gastos = ws["C6"].value or 0
    saldo_disponible = ws["D6"].value or (monto_inicial - total_gastos)
    wb.close()
    return {"monto_inicial": float(monto_inicial), "total_gastos": float(total_gastos), "saldo_disponible": float(saldo_disponible)}

def get_registered_expenses():
    success, msg, mi, tg, sd, expenses = sync_from_google_sheets()
    if success and expenses is not None:
        return expenses

    asegurar_excel_existente()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Control de Gastos"]
    expenses_list = []
    for r in range(11, 60):
        monto = ws[f"F{r}"].value
        if monto is not None and monto != "":
            expenses_list.append({
                "num": ws[f"B{r}"].value,
                "fecha": str(ws[f"C{r}"].value or ""),
                "categoria": ws[f"D{r}"].value or "Otros",
                "descripcion": ws[f"E{r}"].value or "Gasto",
                "monto": float(monto),
                "metodo_pago": ws[f"G{r}"].value or "Efectivo",
                "row_index": r
            })
    wb.close()
    return expenses_list

def add_expense_to_excel(fecha, categoria, descripcion, monto, metodo_pago):
    asegurar_excel_existente()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Control de Gastos"]
    
    target_row = None
    for r in range(11, 60):
        val = ws[f"F{r}"].value
        if val is None or val == "":
            target_row = r
            break
            
    if target_row is None:
        return False, "La tabla de gastos está llena."
        
    ws[f"B{target_row}"] = target_row - 10
    ws[f"C{target_row}"] = fecha
    ws[f"D{target_row}"] = categoria
    ws[f"E{target_row}"] = descripcion
    ws[f"F{target_row}"] = float(monto)
    ws[f"G{target_row}"] = metodo_pago

    wb.save(EXCEL_PATH)
    wb.close()

    google_url = get_google_webhook_url()
    if google_url:
        try:
            requests.post(google_url, json={
                "action": "add_expense",
                "fecha": fecha,
                "categoria": categoria,
                "descripcion": descripcion,
                "monto": float(monto),
                "metodo_pago": metodo_pago
            }, timeout=5)
        except Exception:
            pass

    return True, f"Gasto de ${monto:.2f} registrado con éxito."

def delete_expense_from_excel(row_index, monto, descripcion):
    asegurar_excel_existente()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Control de Gastos"]
    
    deleted = False
    target = int(row_index) if row_index and str(row_index).isdigit() else None
    
    if target and target >= 11 and target < 60:
        ws[f"C{target}"] = None
        ws[f"D{target}"] = None
        ws[f"E{target}"] = None
        ws[f"F{target}"] = None
        ws[f"G{target}"] = None
        deleted = True
    else:
        for r in range(11, 60):
            m = ws[f"F{r}"].value
            d = ws[f"E{r}"].value
            if m is not None and abs(float(m) - float(monto)) < 0.01 and (not descripcion or d == descripcion):
                ws[f"C{r}"] = None
                ws[f"D{r}"] = None
                ws[f"E{r}"] = None
                ws[f"F{r}"] = None
                ws[f"G{r}"] = None
                target = r
                deleted = True
                break

    wb.save(EXCEL_PATH)
    wb.close()

    google_url = get_google_webhook_url()
    if google_url:
        try:
            requests.post(google_url, json={
                "action": "delete_expense",
                "row_index": target,
                "monto": float(monto),
                "descripcion": descripcion
            }, timeout=5)
        except Exception:
            pass

    return deleted

def update_initial_budget(nuevo_monto):
    asegurar_excel_existente()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Control de Gastos"]
    ws["B6"] = float(nuevo_monto)
    wb.save(EXCEL_PATH)
    wb.close()

    google_url = get_google_webhook_url()
    if google_url:
        try:
            requests.post(google_url, json={
                "action": "update_budget",
                "monto_inicial": float(nuevo_monto)
            }, timeout=5)
        except Exception:
            pass

    return True

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Control de Gastos por Voz + Google Sheets</title>
  <script src="https://www.gstatic.com/antigravity/web/dev/tailwindcss.min.js"></script>
  <style>
    .pulse-ring {
      animation: pulse 1.5s infinite;
    }
    @keyframes pulse {
      0% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(239, 68, 68, 0.7); }
      70% { transform: scale(1.05); box-shadow: 0 0 0 15px rgba(239, 68, 68, 0); }
      100% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(239, 68, 68, 0); }
    }
  </style>
</head>
<body class="bg-slate-900 text-slate-100 font-sans antialiased min-h-screen p-3 flex flex-col items-center">

  <div class="w-full max-w-md bg-slate-800 rounded-2xl border border-slate-700 shadow-xl overflow-hidden flex flex-col my-auto">
    
    <div class="bg-indigo-700 p-4 text-center relative flex items-center justify-between">
      <div class="text-left">
        <h1 class="text-lg font-bold text-white flex items-center gap-1.5">
          <span>🎙️</span> Control de Gastos 24/7
        </h1>
        <p id="syncStatus" class="text-[10px] text-indigo-200">🟢 Sincronización Bidireccional</p>
      </div>
      <a href="/download" class="bg-emerald-600 hover:bg-emerald-500 text-white text-xs font-bold px-3 py-1.5 rounded-lg border border-emerald-400 shadow flex items-center gap-1 transition-all">
        📥 Excel (.xlsx)
      </a>
    </div>

    <!-- INTEGRACIÓN GOOGLE SHEETS -->
    <div class="p-3 bg-slate-900/90 border-b border-slate-700 space-y-2">
      <details class="group">
        <summary class="text-xs font-bold text-indigo-300 cursor-pointer flex items-center justify-between">
          <span>📊 Sincronizar con Google Sheets (Opcional)</span>
          <span class="group-open:rotate-180 transition-transform">▼</span>
        </summary>
        <div class="mt-2 space-y-2 text-xs">
          <p class="text-[10px] text-slate-400">Pega aquí la URL de tu Apps Script de Google Sheets para sincronización bidireccional en tiempo real:</p>
          <div class="flex gap-1">
            <input type="text" id="googleUrlInput" class="flex-1 text-[11px] bg-slate-800 border border-slate-700 rounded px-2 py-1 text-slate-200 focus:outline-none" placeholder="https://script.google.com/macros/s/.../exec">
            <button onclick="saveGoogleUrl()" class="bg-emerald-600 hover:bg-emerald-500 text-white font-bold px-2 py-1 rounded text-xs">Vincular</button>
          </div>
        </div>
      </details>

      <div class="flex items-center justify-between gap-2 bg-slate-900 p-2 rounded-xl border border-slate-700">
        <label class="text-xs text-slate-300 font-semibold flex items-center gap-1">
          <span>⚙️</span> Monto Inicial ($):
        </label>
        <div class="flex gap-1">
          <input type="number" step="0.01" id="inputInitialBudget" class="w-24 text-sm font-bold bg-slate-800 text-amber-400 border border-slate-700 rounded-lg px-2 py-1 text-right focus:outline-none focus:border-indigo-500">
          <button onclick="saveInitialBudget()" class="bg-indigo-600 hover:bg-indigo-500 text-white text-xs font-bold px-2 py-1 rounded-lg transition-colors">Guardar</button>
        </div>
      </div>

      <div class="grid grid-cols-3 gap-2">
        <div class="bg-slate-900 p-2 rounded-xl border border-slate-700 text-center">
          <span class="block text-[9px] text-slate-400 font-bold uppercase">Inicial</span>
          <span id="kpiInicial" class="text-xs font-extrabold text-blue-400">$0.00</span>
        </div>
        <div class="bg-slate-900 p-2 rounded-xl border border-slate-700 text-center">
          <span class="block text-[9px] text-slate-400 font-bold uppercase">Gastado</span>
          <span id="kpiGastos" class="text-xs font-extrabold text-rose-400">$0.00</span>
        </div>
        <div class="bg-slate-900 p-2 rounded-xl border border-slate-700 text-center">
          <span class="block text-[9px] text-slate-400 font-bold uppercase">Disponible</span>
          <span id="kpiDisponible" class="text-xs font-extrabold text-emerald-400">$0.00</span>
        </div>
      </div>
    </div>

    <!-- BOTÓN MICRÓFONO -->
    <div class="p-4 flex flex-col items-center justify-center text-center">
      <p class="text-xs text-slate-400 mb-2">Toca el micrófono y di tu gasto libremente:<br><span class="italic text-indigo-300">"Gasté 250 pesos en supermercado comprando fruta y carne"</span></p>

      <button id="btnVoice" onclick="toggleVoice()" class="w-16 h-16 rounded-full bg-indigo-600 hover:bg-indigo-500 text-white flex items-center justify-center text-2xl shadow-lg transition-all transform active:scale-95">
        <span id="micIcon">🎙️</span>
      </button>

      <div id="statusText" class="mt-2 text-xs font-semibold text-slate-300">Toca para dictar por voz</div>
    </div>

    <!-- FORMULARIO DE ENTRADA -->
    <div class="p-4 bg-slate-900/60 border-t border-slate-700 space-y-3">
      <div>
        <label class="block text-[11px] text-slate-400 font-semibold mb-1">🗣️ Texto Dictado o Reconocido:</label>
        <input type="text" id="rawText" class="w-full text-xs bg-slate-800 border border-slate-700 rounded-lg px-3 py-2 text-slate-200 focus:outline-none focus:border-indigo-500" placeholder="Escribe o dicta aquí sin límite de longitud..." oninput="processManualInput()">
      </div>

      <div class="grid grid-cols-2 gap-2">
        <div>
          <label class="block text-[11px] text-slate-400">Monto ($):</label>
          <input type="number" step="0.01" id="valMonto" class="w-full text-sm font-bold bg-slate-800 border border-slate-700 rounded-lg px-3 py-2 text-emerald-400 focus:outline-none">
        </div>
        <div>
          <label class="block text-[11px] text-slate-400">Categoría:</label>
          <select id="valCategoria" class="w-full text-xs bg-slate-800 border border-slate-700 rounded-lg px-2 py-2 text-slate-200 focus:outline-none">
            <option value="Alimentación">Alimentación</option>
            <option value="Vivienda / Servicios">Vivienda / Servicios</option>
            <option value="Transporte">Transporte</option>
            <option value="Entretenimiento">Entretenimiento</option>
            <option value="Salud">Salud</option>
            <option value="Educación">Educación</option>
            <option value="Ropa y Calzado">Ropa y Calzado</option>
            <option value="Otros">Otros</option>
          </select>
        </div>
      </div>

      <div>
        <label class="block text-[11px] text-slate-400 font-semibold mb-1">Descripción / Concepto (Detallado):</label>
        <textarea id="valDesc" rows="2" class="w-full text-xs bg-slate-800 border border-slate-700 rounded-lg px-3 py-2 text-slate-200 focus:outline-none focus:border-indigo-500 resize-none" placeholder="Descripción extendida del gasto..."></textarea>
      </div>

      <div class="grid grid-cols-2 gap-2">
        <div>
          <label class="block text-[11px] text-slate-400">Método Pago:</label>
          <select id="valMetodo" class="w-full text-xs bg-slate-800 border border-slate-700 rounded-lg px-2 py-2 text-slate-200 focus:outline-none">
            <option value="Efectivo">Efectivo</option>
            <option value="Tarjeta de Débito">Tarjeta de Débito</option>
            <option value="Tarjeta de Crédito">Tarjeta de Crédito</option>
            <option value="Transferencia">Transferencia</option>
          </select>
        </div>
        <div class="flex items-end">
          <button onclick="saveExpense()" class="w-full bg-emerald-600 hover:bg-emerald-500 text-white font-bold py-2 rounded-xl text-xs shadow transition-colors">
            ✅ Guardar Gasto
          </button>
        </div>
      </div>

      <!-- TABLA DE HISTORIAL DE GASTOS -->
      <div class="pt-2">
        <div class="flex items-center justify-between mb-2">
          <h3 class="text-xs font-bold uppercase tracking-wider text-slate-300">📋 Gastos en Google Sheets</h3>
          <button onclick="triggerSync()" class="bg-indigo-600 hover:bg-indigo-500 text-white text-[10px] font-bold px-2.5 py-1 rounded-lg flex items-center gap-1 transition-all shadow">
            🔄 Sincronizar en Vivo
          </button>
        </div>
        <div class="overflow-x-auto max-h-48 border border-slate-700 rounded-lg">
          <table class="w-full text-xs text-left text-slate-300">
            <thead class="bg-slate-800 uppercase text-[9px] text-slate-400 sticky top-0">
              <tr>
                <th class="px-2 py-1">Fecha</th>
                <th class="px-2 py-1">Cat.</th>
                <th class="px-2 py-1">Concepto</th>
                <th class="px-2 py-1 text-right">Monto</th>
                <th class="px-1 py-1 text-center">Acción</th>
              </tr>
            </thead>
            <tbody id="historyBody" class="divide-y divide-slate-700/50">
              <tr><td colspan="5" class="px-2 py-2 text-center text-slate-500">Cargando datos...</td></tr>
            </tbody>
          </table>
        </div>
      </div>

      <div id="toast" class="hidden p-2.5 rounded-lg text-xs font-semibold text-center"></div>
    </div>

  </div>

  <script>
    let recognition = null;
    let isListening = false;

    if ('webkitSpeechRecognition' in window || 'SpeechRecognition' in window) {
      const SR = window.SpeechRecognition || window.webkitSpeechRecognition;
      recognition = new SR();
      recognition.lang = 'es-ES';
      recognition.continuous = false;
      recognition.interimResults = false;

      recognition.onstart = function() {
        isListening = true;
        document.getElementById('btnVoice').classList.add('bg-rose-600', 'pulse-ring');
        document.getElementById('btnVoice').classList.remove('bg-indigo-600');
        document.getElementById('micIcon').textContent = '🔴';
        document.getElementById('statusText').textContent = 'Escuchando... habla ahora';
      };

      recognition.onresult = function(event) {
        const transcript = event.results[0][0].transcript;
        document.getElementById('rawText').value = transcript;
        parseSpeechText(transcript);
      };

      recognition.onerror = function(event) {
        showToast('Error de micrófono: ' + event.error, 'error');
        resetMicUI();
      };

      recognition.onend = resetMicUI;
    }

    function resetMicUI() {
      isListening = false;
      document.getElementById('btnVoice').classList.remove('bg-rose-600', 'pulse-ring');
      document.getElementById('micIcon').textContent = '🎙️';
      document.getElementById('statusText').textContent = 'Toca para dictar por voz';
    }

    function toggleVoice() {
      if (!recognition) {
        document.getElementById('rawText').focus();
        return;
      }
      isListening ? recognition.stop() : recognition.start();
    }

    async function parseSpeechText(text) {
      const resp = await fetch('/api/parse_voice', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({text: text})
      });
      const data = await resp.json();
      
      document.getElementById('valMonto').value = data.monto;
      document.getElementById('valCategoria').value = data.categoria;
      document.getElementById('valDesc').value = data.descripcion;
      document.getElementById('valMetodo').value = data.metodo_pago;
    }

    function processManualInput() {
      const text = document.getElementById('rawText').value;
      if (text.length > 2) parseSpeechText(text);
    }

    function fmt(n) { return '$' + Number(n).toLocaleString('en-US', {minimumFractionDigits: 2, maximumFractionDigits: 2}); }

    async function triggerSync() {
      showToast('🔄 Sincronizando con Google Sheets...', 'info');
      const resp = await fetch('/api/sync_google');
      const data = await resp.json();
      if (data.success) {
        showToast('🟢 ' + data.message, 'success');
      } else {
        showToast('⚠️ ' + data.message, 'error');
      }
      loadData();
    }

    async function loadData() {
      const respKpi = await fetch('/api/summary');
      const dataKpi = await respKpi.json();
      document.getElementById('inputInitialBudget').value = dataKpi.monto_inicial;
      document.getElementById('kpiInicial').textContent = fmt(dataKpi.monto_inicial);
      document.getElementById('kpiGastos').textContent = fmt(dataKpi.total_gastos);
      document.getElementById('kpiDisponible').textContent = fmt(dataKpi.saldo_disponible);

      const respGoogle = await fetch('/api/get_google_config');
      const dataGoogle = await respGoogle.json();
      if (dataGoogle.url) {
        document.getElementById('googleUrlInput').value = dataGoogle.url;
        document.getElementById('syncStatus').textContent = '🟢 Google Sheets Vinculado';
      }

      const respHistory = await fetch('/api/expenses');
      const expenses = await respHistory.json();
      const tbody = document.getElementById('historyBody');
      tbody.innerHTML = '';

      if (!expenses || expenses.length === 0) {
        tbody.innerHTML = '<tr><td colspan="5" class="px-2 py-2 text-center text-slate-500">No hay gastos registrados aún en la hoja.</td></tr>';
        return;
      }

      expenses.forEach((item, idx) => {
        const tr = document.createElement('tr');
        tr.className = 'border-b border-slate-700/40';
        tr.innerHTML = `
          <td class="px-2 py-1.5 text-[10px] text-slate-400">${item.fecha}</td>
          <td class="px-2 py-1.5 text-slate-200 font-medium">${item.categoria}</td>
          <td class="px-2 py-1.5 text-slate-400 whitespace-normal break-words max-w-[120px]">${item.descripcion}</td>
          <td class="px-2 py-1.5 text-right text-rose-400 font-bold">${fmt(item.monto)}</td>
          <td class="px-1 py-1.5 text-center">
            <button onclick="deleteExpense(${item.row_index || item.rowIndex || (idx + 11)}, ${item.monto}, '${(item.descripcion || '').replace(/'/g, "\\'")}')" class="bg-rose-900/60 hover:bg-rose-600 text-rose-200 hover:text-white px-2 py-0.5 rounded font-bold text-[10px] transition-colors" title="Eliminar registro">
              🗑️
            </button>
          </td>
        `;
        tbody.appendChild(tr);
      });
    }

    async function deleteExpense(rowIndex, monto, descripcion) {
      if (!confirm(`¿Deseas eliminar este gasto de ${fmt(monto)}?`)) return;

      const resp = await fetch('/api/delete_expense', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({row_index: rowIndex, monto: monto, descripcion: descripcion})
      });
      const data = await resp.json();
      if (data.success) {
        showToast('🗑️ Registro eliminado correctamente', 'success');
        loadData();
      } else {
        showToast('❌ No se pudo eliminar el registro', 'error');
      }
    }

    async function saveGoogleUrl() {
      const url = document.getElementById('googleUrlInput').value.trim();
      const resp = await fetch('/api/set_google_config', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({url: url})
      });
      const data = await resp.json();
      if (data.success) {
        showToast('🟢 Google Sheets vinculado con éxito', 'success');
        document.getElementById('syncStatus').textContent = '🟢 Google Sheets Vinculado';
        triggerSync();
      }
    }

    async function saveInitialBudget() {
      const monto = parseFloat(document.getElementById('inputInitialBudget').value);
      if (isNaN(monto) || monto < 0) {
        showToast('Ingresa un monto inicial válido.', 'error');
        return;
      }
      const resp = await fetch('/api/set_initial_budget', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({monto_inicial: monto})
      });
      const data = await resp.json();
      if (data.success) {
        showToast('⚙️ ' + data.message, 'success');
        loadData();
      }
    }

    async function saveExpense() {
      const monto = parseFloat(document.getElementById('valMonto').value);
      const categoria = document.getElementById('valCategoria').value;
      const descripcion = document.getElementById('valDesc').value;
      const metodo_pago = document.getElementById('valMetodo').value;
      const fecha = new Date().toISOString().split('T')[0];

      if (!monto || monto <= 0) {
        showToast('Ingresa un monto válido.', 'error');
        return;
      }

      const resp = await fetch('/api/add_expense', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({fecha, categoria, descripcion, monto, metodo_pago})
      });
      
      const data = await resp.json();
      if (data.success) {
        showToast('🎉 Gasto guardado correctamente', 'success');
        document.getElementById('rawText').value = '';
        document.getElementById('valMonto').value = '';
        document.getElementById('valDesc').value = '';
        loadData();
      } else {
        showToast('❌ ' + data.message, 'error');
      }
    }

    function showToast(msg, type) {
      const toast = document.getElementById('toast');
      toast.classList.remove('hidden', 'bg-emerald-900', 'text-emerald-200', 'bg-rose-900', 'text-rose-200', 'bg-blue-900', 'text-blue-200');
      if (type === 'success') {
        toast.classList.add('bg-emerald-900', 'text-emerald-200');
      } else if (type === 'error') {
        toast.classList.add('bg-rose-900', 'text-rose-200');
      } else if (type === 'info') {
        toast.classList.add('bg-blue-900', 'text-blue-200');
      }
      toast.textContent = msg;
      setTimeout(() => { toast.classList.add('hidden'); }, 5000);
    }

    loadData();
  </script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route("/api/summary", methods=["GET"])
def summary():
    return jsonify(get_excel_summary())

@app.route("/api/expenses", methods=["GET"])
def expenses():
    return jsonify(get_registered_expenses())

@app.route("/api/sync_google", methods=["GET"])
def sync_google():
    success, msg, mi, tg, sd, exp = sync_from_google_sheets()
    return jsonify({"success": success, "message": msg, "total_synced": len(exp) if exp else 0})

@app.route("/api/parse_voice", methods=["POST"])
def parse_voice():
    data = request.get_json() or {}
    text = data.get("text", "")
    return jsonify(parse_voice_text(text))

@app.route("/api/add_expense", methods=["POST"])
def add_expense():
    data = request.get_json() or {}
    fecha = data.get("fecha", datetime.now().strftime("%Y-%m-%d"))
    categoria = data.get("categoria", "Otros")
    descripcion = data.get("descripcion", "Gasto")
    monto = data.get("monto", 0)
    metodo_pago = data.get("metodo_pago", "Efectivo")
    
    success, msg = add_expense_to_excel(fecha, categoria, descripcion, monto, metodo_pago)
    return jsonify({"success": success, "message": msg})

@app.route("/api/delete_expense", methods=["POST"])
def delete_expense():
    data = request.get_json() or {}
    row_index = data.get("row_index")
    monto = data.get("monto", 0)
    descripcion = data.get("descripcion", "")
    
    success = delete_expense_from_excel(row_index, monto, descripcion)
    return jsonify({"success": success})

@app.route("/api/set_initial_budget", methods=["POST"])
def set_initial_budget():
    data = request.get_json() or {}
    monto_inicial = data.get("monto_inicial", 5000)
    update_initial_budget(monto_inicial)
    return jsonify({"success": True, "message": f"Monto inicial modificado a ${float(monto_inicial):.2f}"})

@app.route("/api/get_google_config", methods=["GET"])
def get_google_config():
    return jsonify({"url": get_google_webhook_url()})

@app.route("/api/set_google_config", methods=["POST"])
def set_google_config():
    data = request.get_json() or {}
    url = data.get("url", "").strip()
    set_google_webhook_url(url)
    return jsonify({"success": True, "url": url})

@app.route("/download")
def download():
    asegurar_excel_existente()
    return send_file(EXCEL_PATH, as_attachment=True, download_name="Control_de_Gastos_Actualizado.xlsx")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
